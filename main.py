import numpy as np
import pandas as pd
import constants
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils.cell import get_column_letter


def clean_branch_terr_column(df):
    """O(n)"""
    df['Branch/Terr.'] = df['Branch/Terr.'].fillna('ALL')


def insert_rate_card_dupe_check_column(df):
    """O(n)"""
    df.insert(1, "Dupe Check", df['Branch/Terr.'] + ':' + df['Customer Class'] + ':' + df['PROD_NBR'])


def insert_rate_card_key_column(df):
    """O(n)"""
    df.insert(1, "Key", df['Customer Class'] + ':' + df['PROD_NBR'])


def clean_prod_nbr_column(df):
    """O(n)"""
    df.loc[df['PROD_NBR'].str.startswith('G'), 'PROD_NBR'] = df['PROD_NBR'].str[1:]


def insert_ord_comm_column(df):
    """O(n)"""
    std_cost_column = df.columns.get_loc('STD-COST')
    ord_comm_column = df.columns.get_loc('AVG-COST')
    df.insert(ord_comm_column, 'Ord Comm', '')
    df['Ord Comm'] = f'=INDIRECT("RC[{std_cost_column - ord_comm_column}]",0)'


def insert_special_dupe_check_column(df):
    """O(n)"""
    dupe_column = df.columns.get_loc('Matrix Tag Comment')
    df.insert(dupe_column, "Dupe Check", df['Branch/Terr.'] + ':' + df['Customer ID'] + ':' + df['PROD_NBR'])


def insert_special_key_column(df):
    """O(n)"""
    key_column = df.columns.get_loc('Matrix Tag Comment')
    df.insert(key_column, "Key", df['Customer ID'] + ':' + df['PROD_NBR'])


def get_basis_name(n):
    """O(1)"""
    basis_names = {'1': 'LIST', '2': 'INTERNAL', '3': 'UMSP', '4': 'CMP', '5': 'STD-COST',
                   '6': 'REP-COST', '8': 'AVG-COST', '9': 'LASTCOST', '10': 'CMP',
                   '21': 'Lnd Cost', '22': 'Avg Lnd', '25': 'Ord COGS', '28': 'Ord Comm',
                   '29': 'Strgc List', '30': 'Strgc Cost', '31': 'AVG-COST'}
    return basis_names.get(n, n)


def insert_job_column(df):
    """O(n)"""
    job_column = df.columns.get_loc('Bill To ID')
    df.insert(job_column, "Job", '')
    df['Job'] = np.where(df['Bill To ID'].isnull(), "No", "Yes")


def insert_basis_names(df):
    """O(n^4)?"""
    # insert the price basis column
    price_formula_column = df.columns.get_loc('Price Formula')
    df.insert(price_formula_column, 'Price Basis', '')
    df['Price Basis'] = df['Price Basis#'].apply(get_basis_name)

    # insert the cost basis column
    price_formula_column = df.columns.get_loc('Cost Formula')
    df.insert(price_formula_column, 'Cost Basis', '')
    df['Cost Basis'] = df['Cost Basis#'].apply(get_basis_name)


def insert_price_multiplier(df):
    """O(n)"""
    price_multiplier_column = df.columns.get_loc('Qty Break Basis')
    df.insert(price_multiplier_column, "Price Multiplier", """=_xlfn.IFERROR(_xlfn.SWITCH(LEFT(_xlfn.INDIRECT("RC[-1]", 0), 1),
        "*", VALUE(MID(_xlfn.INDIRECT("RC[-1]", 0), 2, LEN(_xlfn.INDIRECT("RC[-1]", 0)))),
        "X", VALUE(MID(_xlfn.INDIRECT("RC[-1]", 0), 2, LEN(_xlfn.INDIRECT("RC[-1]", 0)))),
        "+", 1 + VALUE(_xlfn.INDIRECT("RC[-1]", 0)) / 100,
        "-", 1 + VALUE(_xlfn.INDIRECT("RC[-1]", 0)) / 100,
        "D", 1 / VALUE(MID(_xlfn.INDIRECT("RC[-1]", 0), 2, LEN(_xlfn.INDIRECT("RC[-1]", 0)))),
        "G", 1 / (1 - VALUE(MID(_xlfn.INDIRECT("RC[-1]", 0), 3, LEN(_xlfn.INDIRECT("RC[-1]", 0))) / 100)),
        "$",0),"")"""
              )


def insert_formula_price_basis(df):
    """O(n)"""
    formula_price_basis_column = df.columns.get_loc('Price Basis')
    df.insert(formula_price_basis_column, 'Hard Price Basis', df['Price Basis'])
    df['Price Basis'] = '=@INDIRECT("RC[-2]", 0)'


def insert_formula_price_formula(df):
    formula_price_basis_column = df.columns.get_loc('Price Basis')
    df.insert(formula_price_basis_column, 'Hard Price Formula', df['Price Formula'])
    df['Price Formula'] = '=@INDIRECT("RC[-2]", 0)'


def insert_unit_price(df):
    unit_price_column = df.columns.get_loc('Qty Break Basis')
    df.insert(unit_price_column, 'Unit Price', """=IF(_xlfn.INDIRECT("RC[-1]", 0)=0, VALUE(_xlfn.INDIRECT("RC[-2]", 0)), _xlfn.INDIRECT("RC[-1]", 0)*INDEX(_xlfn.INDIRECT("R[0]C[-16]:R[0]C[-7]",0),1,MATCH(_xlfn.INDIRECT("RC[-3]",0),_xlfn.INDIRECT("R1C[-16]:R1C[-7]",0),0
    )))""")


def insert_cost_gp_perc(df):
    cost_gp_perc_column = df.columns.get_loc('Qty Break Basis')
    df.insert(cost_gp_perc_column, "Cost GP %", '=1-_xlfn.indirect("RC[-12]",0)/_xlfn.indirect("RC[-1]",0)')


def insert_cogs_gp_perc(df):
    cogs_gp_perc_column = df.columns.get_loc('Qty Break Basis')
    df.insert(cogs_gp_perc_column, "CoGS GP %", '=1-_xlfn.indirect("RC[-11]",0)/_xlfn.indirect("RC[-2]",0)')


def insert_ord_comm_basis_column(df):
    """O(n)"""
    price_multiplier_column = df.columns.get_loc('Effective Date')
    df.insert(price_multiplier_column, "Ord Comm Basis", "")


def insert_ord_comm_formula_column(df):
    """O(n)"""
    price_multiplier_column = df.columns.get_loc('Effective Date')
    df.insert(price_multiplier_column, "Ord Comm Formula", "")


def insert_ord_comm_multiplier_column(df):
    """O(n)"""
    price_multiplier_column = df.columns.get_loc('Effective Date')
    df.insert(price_multiplier_column, "Ord Comm Multiplier", """=_xlfn.IFERROR(_xlfn.SWITCH(LEFT(_xlfn.INDIRECT("RC[-1]", 0), 1),
     "*", VALUE(MID(_xlfn.INDIRECT("RC[-1]", 0), 2, LEN(_xlfn.INDIRECT("RC[-1]", 0)))),
     "X", VALUE(MID(_xlfn.INDIRECT("RC[-1]", 0), 2, LEN(_xlfn.INDIRECT("RC[-1]", 0)))),
     "+", 1 + VALUE(_xlfn.INDIRECT("RC[-1]", 0)) / 100,
     "-", 1 + VALUE(_xlfn.INDIRECT("RC[-1]", 0)) / 100,
     "D", 1 / VALUE(MID(_xlfn.INDIRECT("RC[-1]", 0), 2, LEN(_xlfn.INDIRECT("RC[-1]", 0)))),
     "G", 1 / (1 - VALUE(MID(_xlfn.INDIRECT("RC[-1]", 0), 3, LEN(_xlfn.INDIRECT("RC[-1]", 0))) / 100)),
     "$",0),"")""")


def insert_ord_comm_value_column(df):
    """O(n)"""
    price_multiplier_column = df.columns.get_loc('Effective Date')
    df.insert(price_multiplier_column, "Ord Comm Value", "")


def split_branches(branches, cs, bs, rc, hl, branch_list):
    # for branch in cs['Home Branch'].unique():
    for branch in branch_list:
        hierarchy = hl[branch]
        branches[branch] = {
            'cs_price': cs.loc[(~cs['Price Basis#'].isnull()) &
                               (cs['Home Branch'] == branch)],
            'cs_cost_part': cs.loc[(~cs['Cost Basis#'].isnull()) &
                                   (cs['Home Branch'] == branch) &
                                   (cs['PROD_NBR'].str.isdecimal())],
            'cs_cost_group': cs.loc[(~cs['Cost Basis#'].isnull()) &
                                    (cs['Home Branch'] == branch) &
                                    (~cs['PROD_NBR'].str.isdecimal())],
            'bs_price': bs.loc[(~bs['Price Basis#'].isnull()) &
                               (bs['Branch/Terr.'] == branch)],
            'bs_cost_part': bs.loc[(~bs['Cost Basis#'].isnull()) &
                                   (bs['Branch/Terr.'] == branch) &
                                   (bs['PROD_NBR'].str.isdecimal())],
            'bs_cost_group': bs.loc[(~bs['Cost Basis#'].isnull()) &
                                    (bs['Branch/Terr.'] == branch) &
                                    (~bs['PROD_NBR'].str.isdecimal())],
            'rc_price_part': rc.loc[(~rc['Price Basis#'].isnull()) &
                                    (rc['Branch/Terr.'].isin(hierarchy)) &
                                    (rc['PROD_NBR'].str.isdecimal())],
            'rc_price_group': rc.loc[(~rc['Price Basis#'].isnull()) &
                                     (rc['Branch/Terr.'].isin(hierarchy)) &
                                     (~rc['PROD_NBR'].str.isdecimal())],
            'rc_cost_part': rc.loc[(~rc['Cost Basis#'].isnull()) &
                                   (rc['Branch/Terr.'].isin(hierarchy)) &
                                   (rc['PROD_NBR'].str.isdecimal())],
            'rc_cost_group': rc.loc[(~rc['Cost Basis#'].isnull()) &
                                    (rc['Branch/Terr.'].isin(hierarchy)) &
                                    (~rc['PROD_NBR'].str.isdecimal())],
        }


def write_excel_workbooks(branches):
    for branch in branches:
        print(branch)
        wb = openpyxl.Workbook()
        for key in branches[branch]:
            ws = wb.create_sheet(key)
            for r in dataframe_to_rows(branches[branch][key], index=False, header=True):
                ws.append(r)
        wb.save(branch + ".xlsx")
        #
        # with pd.ExcelWriter(branch + ".xlsx", mode='w') as writer:
        #     # branches[branch]['cs_price'].to_excel(writer, sheet_name='CS-Price')
        #     branches[branch]['cs_cost_part'].to_excel(writer, sheet_name='CS-Cost-Part')
        #     branches[branch]['cs_cost_group'].to_excel(writer, sheet_name='CS-Cost-Group')
        #     branches[branch]['bs_price'].to_excel(writer, sheet_name='BS-Price')
        #     branches[branch]['bs_cost_part'].to_excel(writer, sheet_name='BS-Cost-Part')
        #     branches[branch]['bs_cost_group'].to_excel(writer, sheet_name='BS-Cost-Group')
        #     branches[branch]['rc_price_part'].to_excel(writer, sheet_name='RC-Price-Part')
        #     branches[branch]['rc_price_group'].to_excel(writer, sheet_name='RC-Price-Group')
        #     branches[branch]['rc_cost_part'].to_excel(writer, sheet_name='RC-Cost-Part')
        #     branches[branch]['rc_cost_group'].to_excel(writer, sheet_name='RC-Cost-Group')


def main():
    pd.set_option('display.max_columns', None)

    cs = pd.read_csv('cs.csv', encoding='Windows 1252', dtype=constants.SPECIAL_DATA_TYPES)
    bs = pd.read_csv('bs.csv', encoding='Windows 1252', dtype=constants.SPECIAL_DATA_TYPES)
    rc = pd.read_csv('rc.csv', encoding='Windows 1252', dtype=constants.RATE_CARD_DATA_TYPES)
    hl = pd.read_csv('Hierarchy.csv', encoding='Windows 1252')

    # remove the 'G' from in front of the PROD_NBRs that are groups
    clean_prod_nbr_column(cs)
    clean_prod_nbr_column(bs)
    clean_prod_nbr_column(rc)

    # change NaNs to 'ALL'
    clean_branch_terr_column(cs)
    clean_branch_terr_column(bs)
    clean_branch_terr_column(rc)

    insert_job_column(cs)
    insert_job_column(bs)

    insert_special_dupe_check_column(cs)
    insert_special_dupe_check_column(bs)

    insert_special_key_column(cs)
    insert_special_key_column(bs)

    insert_price_multiplier(cs)
    insert_price_multiplier(bs)

    insert_basis_names(cs)
    insert_basis_names(bs)
    insert_basis_names(rc)

    insert_formula_price_basis(cs)
    insert_formula_price_basis(bs)

    insert_formula_price_formula(cs)
    insert_formula_price_formula(bs)

    insert_unit_price(cs)
    insert_unit_price(bs)

    insert_cost_gp_perc(cs)
    insert_cost_gp_perc(bs)

    insert_cogs_gp_perc(cs)
    insert_cogs_gp_perc(bs)

    insert_ord_comm_column(cs)
    insert_ord_comm_column(bs)

    insert_ord_comm_basis_column(cs)
    insert_ord_comm_formula_column(cs)
    insert_ord_comm_multiplier_column(cs)
    insert_ord_comm_value_column(cs)

    insert_ord_comm_basis_column(bs)
    insert_ord_comm_formula_column(bs)
    insert_ord_comm_multiplier_column(bs)
    insert_ord_comm_value_column(bs)

    insert_rate_card_dupe_check_column(rc)
    insert_rate_card_key_column(rc)

    branches = {}
    branch_list = list(set(cs['Home Branch'].unique()).union(set(bs['Home Branch'].unique())))
    branch_list.sort()
    branch_list.remove('1000')
    split_branches(branches, cs, bs, rc, hl, branch_list)

    write_excel_workbooks(branches)


if __name__ == '__main__':
    main()
